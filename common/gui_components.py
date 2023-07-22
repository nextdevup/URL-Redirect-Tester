import pandas as pd
import re
import tkinter as tk
import tkinter.filedialog
import tkinter.ttk as ttk
from tkinter import messagebox
from openpyxl import load_workbook

class InputRow:
    def __init__(self, master, row, label, btn_method, popup_options):
        self.master = master
        self.frm = ttk.Frame(width=100, height=20)
        self.frm.grid(row=row, column=0, padx=5, pady=5, sticky="w")
        self.lbl = ttk.Label(text=label, width=27, master=self.frm, border=1, relief=tk.SOLID)
        self.lbl.grid(row=0, column=0, sticky="e")
        self.lbl_selection = None

        if btn_method is None and popup_options is None:
            self.ent = ttk.Entry(
                background="white",
                foreground="black",
                master=self.frm,
                width=50
            )
            self.ent.grid(row=0, column=1, sticky="w")
        else:
            self.lbl_selection = ttk.Label(text="", width=50, master=self.frm, border=1, relief=tk.SOLID)
            self.lbl_selection.grid(row=0, column=1, sticky="w")

            if not popup_options is None:
                self.btn = ttk.Button(
                    master=self.frm,
                    text="Select",
                    width=20,
                    command=lambda: self.open_popup(popup_options, self.set_value)
                )
            else:
                self.btn = ttk.Button(
                    master=self.frm,
                    text="Select",
                    width=20,
                    command=lambda: btn_method(self.set_value)
                )

            self.btn.grid(row=0, column=2, sticky="w")
    def get_column_names(self):
        file = pd.read_excel(self.get_selection_value())
        columns = file.columns.values

        #If first row has paths/URLs then there isn't a column name and we should use the column index instead
        if any(col_name.startswith("http") or col_name.startswith("/") for col_name in columns):
            columns = map(str, list(range(len(columns))))

        return columns
    def get_excel_sheet_names(self):
        workbook = load_workbook(self.get_selection_value(), read_only=True, keep_links=False)
        return workbook.sheetnames
    def get_selection_value(self):
        if self.lbl_selection is None:
            return str(self.ent.get()).strip()
        else:
            return str(self.lbl_selection["text"]).strip()
    def open_popup(self, selection_options, btn_method):
        popup = tk.Toplevel()
        popup.wm_title(self.lbl['text'])

        lbl_popup = tk.Label(popup, text=self.lbl['text'], width=50)
        lbl_popup.grid(row=0, column=0)

        self.selection = tk.StringVar(popup.master)
        popup.ddl_popup = ttk.OptionMenu(popup, self.selection, *selection_options())
        popup.ddl_popup.grid(row=1, column=0)
        btn_popup = ttk.Button(popup, text="Select", command=lambda: [btn_method(self.selection.get()), popup.destroy()])
        btn_popup.grid(row=2, column=0)
    def set_value(self, selected_value):
        self.lbl_selection.config(text=str(selected_value).strip())
